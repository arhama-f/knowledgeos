import io

from openpyxl import load_workbook

from app.services.extractors.base import ExtractedPage, Extractor

MAX_ROWS_PER_SHEET = 20_000


class ExcelExtractor(Extractor):
    def extract(self, content: bytes) -> list[ExtractedPage]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        pages = []
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            rows = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= MAX_ROWS_PER_SHEET:
                    break
                cells = [str(c).strip() for c in row if c is not None]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                text = f"Sheet: {sheet.title}\n\n" + "\n\n".join(rows)
                pages.append(ExtractedPage(page_number=sheet_index, text=text))
        return pages
